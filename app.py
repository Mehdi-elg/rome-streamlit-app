
import streamlit as st
import requests
import pandas as pd
import io
import time
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Recherche ROME - FIPU", page_icon="🔎", layout="wide")

CLIENT_ID = st.secrets["CLIENT_ID"]
CLIENT_SECRET = st.secrets["CLIENT_SECRET"]

TOKEN_URL = "https://entreprise.francetravail.fr/connexion/oauth2/access_token?realm=/partenaire"
API_BASE = "https://api.francetravail.io/partenaire/rome-metiers/v1/metiers"
SCOPES = "nomenclatureRome api_rome-metiersv1"

CHAMPS_METIER = "code,libelle,contextesTravail(categorie,libelle)"

CRITERES_FIPU = [
    "En altitude", "En milieu nucléaire", "En milieu hyperbare",
    "En milieu exigu ou confiné", "En grande hauteur", "En zone frigorifique",
    "Exposition à de hautes températures", "En environnement climatique difficile",
    "Manipulation d'un engin, équipement ou outil dangereux",
    "Port et manipulation de charges lourdes ou encombrantes",
    "Position pénible", "Station debout prolongée",
    "Travail répétitif ou cadence imposée", "En environnement bruyant",
    "Travail dans des environnements hostiles et dangereux",
    "Exposition à de basses températures",
    "Exposition possible à gaz, aérosol, fumées …",
    "Station assise prolongée", "Risques de chutes",
    "Travail dans des milieux difficiles et exigeants pour l'humain",
    "Travail posté (2x8, 3x8, 5x8, etc.)", "Travail de nuit",
    "Travail en astreinte", "Travail en horaires décalés",
    "Travail par roulement",
]
CRITERES_FIPU_LOWER = [c.lower() for c in CRITERES_FIPU]

SEPARATEUR = " ; "
CHUNK_SIZE_FULL = 25          # nb de codes traités par "vague" dans l'onglet 2
MAX_RETRIES = 3
TIMEOUT = 15

# ══════════════════════════════════════════════════════════════════════════
# COUCHE API (authentification, appels réseau)
# ══════════════════════════════════════════════════════════════════════════

def get_session() -> requests.Session:
    """Session HTTP réutilisable (connexion garde-alive)."""
    if "http_session" not in st.session_state:
        st.session_state.http_session = requests.Session()
    return st.session_state.http_session


def get_token() -> str:
    """Récupère un token OAuth2 et le met en cache tant qu'il est valide."""
    now = datetime.utcnow()
    token = st.session_state.get("api_token")
    expiry = st.session_state.get("api_token_expiry")
    if token and expiry and now < expiry:
        return token

    session = get_session()
    data = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": SCOPES,
    }
    r = session.post(TOKEN_URL, data=data, timeout=TIMEOUT)
    r.raise_for_status()
    payload = r.json()
    token = payload["access_token"]
    expires_in = payload.get("expires_in", 1500)
    st.session_state.api_token = token
    # marge de sécurité de 60s avant expiration réelle
    st.session_state.api_token_expiry = now + timedelta(seconds=max(expires_in - 60, 30))
    return token


def api_get(url: str, params: dict | None = None) -> requests.Response:
    """GET authentifié avec retry automatique en cas d'échec réseau ou de token expiré."""
    session = get_session()
    last_exc = None
    for attempt in range(MAX_RETRIES):
        token = get_token()
        headers = {"Authorization": f"Bearer {token}"}
        try:
            r = session.get(url, headers=headers, params=params, timeout=TIMEOUT)
            if r.status_code == 401:
                # token invalide -> on force son renouvellement puis on retente
                st.session_state.api_token = None
                continue
            r.raise_for_status()
            return r
        except requests.RequestException as exc:
            last_exc = exc
            time.sleep(0.5 * (attempt + 1))
    if last_exc:
        raise last_exc
    raise RuntimeError("Échec de la requête API après plusieurs tentatives.")


def get_metier(code_rome: str) -> dict:
    url = f"{API_BASE}/metier/{code_rome}"
    r = api_get(url, params={"champs": CHAMPS_METIER})
    return r.json()


def get_all_rome_codes() -> list[str]:
    """Récupère la liste complète des codes ROME exposés par l'API."""
    url = f"{API_BASE}/metier"
    r = api_get(url)
    data = r.json()
    items = data if isinstance(data, list) else data.get("metiers", data.get("results", []))
    codes = []
    for item in items:
        code = item.get("code") if isinstance(item, dict) else item
        if code:
            codes.append(code)
    return codes


# ══════════════════════════════════════════════════════════════════════════
# COUCHE MÉTIER (extraction des libellés + règle FIPU)
# ══════════════════════════════════════════════════════════════════════════

def get_libelles_by_categorie(metier: dict, categorie: str) -> list[str]:
    libelles = []
    for ctx in metier.get("contextesTravail", []):
        if ctx.get("categorie") == categorie:
            libelle = (ctx.get("libelle") or "").strip()
            if libelle:
                libelles.append(libelle)
    return libelles


def is_fipu(conditions: list[str], horaires: list[str]) -> bool:
    """Règle de décision FIPU : au moins un libellé correspond à un critère FIPU."""
    texte = " ".join(conditions + horaires).lower()
    return any(c in texte for c in CRITERES_FIPU_LOWER)


def build_row(metier: dict) -> dict:
    conditions = get_libelles_by_categorie(metier, "CONDITIONS_TRAVAIL")
    horaires = get_libelles_by_categorie(metier, "HORAIRE_ET_DUREE_TRAVAIL")
    return {
        "Code ROME": metier.get("code", ""),
        "Libellé": metier.get("libelle", ""),
        "Conditions de travail et risques professionnels": SEPARATEUR.join(conditions),
        "Horaires et durée du travail": SEPARATEUR.join(horaires),
        "FIPU": "Oui" if is_fipu(conditions, horaires) else "Non",
    }


def process_codes(codes: list[str], progress_bar, status_text, stop_check=None) -> tuple[list[dict], list[dict]]:
    """Interroge l'API pour chaque code, met à jour la barre de progression.
    Retourne (lignes_ok, erreurs)."""
    rows, errors = [], []
    total = len(codes)
    for i, code in enumerate(codes):
        if stop_check and stop_check():
            status_text.warning(f"⛔ Recherche interrompue à {i}/{total} codes traités.")
            break
        try:
            metier = get_metier(code)
            rows.append(build_row(metier))
        except requests.HTTPError as e:
            errors.append({"code": code, "erreur": f"HTTP {e.response.status_code if e.response else '?'}"})
        except Exception as e:
            errors.append({"code": code, "erreur": str(e)[:80]})
        progress_bar.progress((i + 1) / total)
        status_text.text(f"{i + 1} / {total} codes traités…")
    return rows, errors


# ══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
FIPU_OUI_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FIPU_NON_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Resultats_ROME") -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        # En-têtes
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = {"Code ROME": 12, "Libellé": 35,
                  "Conditions de travail et risques professionnels": 60,
                  "Horaires et durée du travail": 50, "FIPU": 10}
        for col_idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = widths.get(col_name, 25)

        fipu_col_idx = list(df.columns).index("FIPU") + 1 if "FIPU" in df.columns else None
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (3, 4)))
            if fipu_col_idx:
                fcell = ws.cell(row=row_idx, column=fipu_col_idx)
                fcell.alignment = Alignment(horizontal="center", vertical="center")
                fcell.fill = FIPU_OUI_FILL if fcell.value == "Oui" else FIPU_NON_FILL

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf.seek(0)
    return buf.getvalue()


def styled_preview(df: pd.DataFrame):
    """Aperçu coloré du tableau dans l'app (FIPU en couleur)."""
    def highlight_fipu(val):
        if val == "Oui":
            return "background-color: #C6EFCE; color: #006100; font-weight: 600;"
        if val == "Non":
            return "background-color: #FFC7CE; color: #9C0006; font-weight: 600;"
        return ""
    styler = df.style.applymap(highlight_fipu, subset=["FIPU"]) if "FIPU" in df.columns else df.style
    st.dataframe(styler, use_container_width=True, height=min(45 + 35 * len(df), 600))


def parse_codes(raw_text: str) -> tuple[list[str], int]:
    seen, codes = set(), []
    lines = [l.strip().upper() for l in raw_text.strip().split("\n") if l.strip()]
    for code in lines:
        if code not in seen:
            seen.add(code)
            codes.append(code)
    return codes, len(lines) - len(codes)


# ══════════════════════════════════════════════════════════════════════════
# INTERFACE
# ══════════════════════════════════════════════════════════════════════════

st.title("🔎 Recherche Métiers ROME — Conditions, Horaires & FIPU")

tab_list, tab_all = st.tabs(["📋 Recherche par codes ROME", "🌐 Extraction complète"])

# ── ONGLET 1 ─────────────────────────────────────────────────────────────
with tab_list:
    st.markdown("Collez une liste de codes ROME (un code par ligne), puis lancez la recherche.")

    codes_input = st.text_area(
        "Codes ROME",
        height=150,
        placeholder="A1413\nM1805\nH1203",
        key="tab1_input",
    )

    launch = st.button("🔍 Lancer la recherche", type="primary", key="tab1_launch")

    if launch:
        # Réinitialisation systématique : l'ancien résultat disparaît immédiatement
        st.session_state.tab1_df = None
        st.session_state.tab1_errors = []
        st.session_state.tab1_excel = None

        codes, doublons = parse_codes(codes_input)
        if not codes:
            st.warning("⚠️ Veuillez entrer au moins un code ROME valide.")
        else:
            if doublons:
                st.info(f"ℹ️ {doublons} doublon(s) ignoré(s).")
            progress_bar = st.progress(0)
            status_text = st.empty()
            with st.spinner("Recherche en cours…"):
                rows, errors = process_codes(codes, progress_bar, status_text)
            status_text.empty()
            progress_bar.empty()

            if rows:
                st.session_state.tab1_df = pd.DataFrame(rows)
                st.session_state.tab1_excel = to_excel_bytes(st.session_state.tab1_df)
            st.session_state.tab1_errors = errors
            st.session_state.tab1_total = len(codes)

    # Affichage des résultats (uniquement si une recherche a réussi)
    if st.session_state.get("tab1_df") is not None:
        df1 = st.session_state.tab1_df
        errors1 = st.session_state.get("tab1_errors", [])
        total1 = st.session_state.get("tab1_total", len(df1))

        c1, c2, c3 = st.columns(3)
        c1.metric("Codes demandés", total1)
        c2.metric("Résultats obtenus", len(df1))
        c3.metric("FIPU = Oui", int((df1["FIPU"] == "Oui").sum()))

        st.subheader("📊 Résultats")
        styled_preview(df1)

        st.download_button(
            label=f"📥 Télécharger les résultats (Excel) — {len(df1)} métiers",
            data=st.session_state.tab1_excel,
            file_name=f"ROME_resultats_{len(df1)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="tab1_download",
        )

        if errors1:
            with st.expander(f"⚠️ {len(errors1)} code(s) en erreur"):
                st.dataframe(pd.DataFrame(errors1), use_container_width=True)

    with st.expander("💡 Exemple d'utilisation"):
        st.code("A1413\nM1805\nH1203\nK2110", language="text")

# ── ONGLET 2 ─────────────────────────────────────────────────────────────
with tab_all:
    st.markdown(
        "Récupère **tous les codes ROME** exposés par l'API puis interroge chaque fiche métier. "
        "L'opération peut prendre plusieurs minutes ; elle peut être interrompue à tout moment via **STOP**."
    )

    defaults = {
        "full_running": False,
        "full_stop": False,
        "full_codes": [],
        "full_index": 0,
        "full_rows": [],
        "full_errors": [],
        "full_df": None,
        "full_excel": None,
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)

    c1, c2 = st.columns(2)
    with c1:
        start_full = st.button(
            "🚀 Lancer la recherche complète",
            type="primary",
            disabled=st.session_state.full_running,
            key="tab2_start",
        )
    with c2:
        stop_full = st.button(
            "⛔ STOP",
            disabled=not st.session_state.full_running,
            key="tab2_stop",
        )

    if start_full:
        st.session_state.full_running = True
        st.session_state.full_stop = False
        st.session_state.full_codes = []
        st.session_state.full_index = 0
        st.session_state.full_rows = []
        st.session_state.full_errors = []
        st.session_state.full_df = None
        st.session_state.full_excel = None
        st.rerun()

    if stop_full:
        st.session_state.full_stop = True

    # Boucle de traitement par petits lots (permet au bouton STOP de réagir)
    if st.session_state.full_running:
        if not st.session_state.full_codes:
            with st.spinner("Récupération de la liste complète des codes ROME…"):
                try:
                    st.session_state.full_codes = get_all_rome_codes()
                except Exception as e:
                    st.error(f"Erreur lors de la récupération des codes ROME : {e}")
                    st.session_state.full_running = False
                    st.session_state.full_codes = []

        codes_all = st.session_state.full_codes
        total = len(codes_all)

        if total:
            start_idx = st.session_state.full_index
            end_idx = min(start_idx + CHUNK_SIZE_FULL, total)

            progress_bar2 = st.progress(start_idx / total)
            status_text2 = st.empty()
            status_text2.text(f"{start_idx} / {total} codes traités…")

            for i in range(start_idx, end_idx):
                if st.session_state.full_stop:
                    break
                code = codes_all[i]
                try:
                    metier = get_metier(code)
                    st.session_state.full_rows.append(build_row(metier))
                except requests.HTTPError as e:
                    st.session_state.full_errors.append(
                        {"code": code, "erreur": f"HTTP {e.response.status_code if e.response else '?'}"}
                    )
                except Exception as e:
                    st.session_state.full_errors.append({"code": code, "erreur": str(e)[:80]})
                st.session_state.full_index = i + 1
                progress_bar2.progress((i + 1) / total)
                status_text2.text(f"{i + 1} / {total} codes traités…")

            if st.session_state.full_stop or st.session_state.full_index >= total:
                st.session_state.full_running = False
                rows = st.session_state.full_rows
                st.session_state.full_df = pd.DataFrame(rows) if rows else pd.DataFrame()
                if rows:
                    st.session_state.full_excel = to_excel_bytes(
                        st.session_state.full_df, sheet_name="Tous_ROME_FIPU"
                    )
                st.rerun()
            else:
                time.sleep(0.05)
                st.rerun()

    # Affichage des résultats persistants (pendant et après la recherche)
    df_all = st.session_state.get("full_df")
    if df_all is not None:
        if df_all.empty:
            st.warning("Aucun métier récupéré.")
        else:
            nb_ok = len(df_all)
            nb_err = len(st.session_state.full_errors)
            nb_fipu = int((df_all["FIPU"] == "Oui").sum())
            statut = "⛔ Recherche interrompue" if st.session_state.full_stop else "✅ Recherche terminée"
            st.success(f"{statut} — {nb_ok} métiers récupérés, {nb_err} erreurs.")

            ca, cb, cc = st.columns(3)
            ca.metric("Métiers récupérés", nb_ok)
            cb.metric("Erreurs / non trouvés", nb_err)
            cc.metric("Métiers FIPU = Oui", nb_fipu)

            st.subheader("📊 Résultats")
            styled_preview(df_all)

            st.download_button(
                label=f"📥 Télécharger les résultats (Excel) — {nb_ok} métiers",
                data=st.session_state.full_excel,
                file_name=f"ROME_extraction_complete_{nb_ok}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="tab2_download",
            )

            if nb_err > 0:
                with st.expander(f"⚠️ {nb_err} code(s) en erreur"):
                    st.dataframe(pd.DataFrame(st.session_state.full_errors), use_container_width=True)
